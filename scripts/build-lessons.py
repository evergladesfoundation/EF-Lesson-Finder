#!/usr/bin/env python3
"""Compile Active rows from the Master Lesson Index into widget/src/data/lessons.ts."""

from __future__ import annotations

import json
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / "data" / "Everglades_Master_Lesson_Index.xlsx"
PDF_META = ROOT / "data" / "pdf-metadata.json"
OUT = ROOT / "widget" / "src" / "data" / "lessons.ts"

CONCEPTS = {
    1: "The Everglades is unique and valuable.",
    2: "The Everglades is defined and connected by water.",
    3: "The Everglades is shaped by southern Florida's geology and geography.",
    4: "The Everglades influences and is influenced by weather and climate.",
    5: "The Everglades supports and is connected by a great diversity of life and ecosystems.",
    6: "The Everglades has experienced many changes over time and is endangered.",
    7: "The Everglades and people are inextricably interconnected.",
}


def split_list(raw: object) -> list[str]:
    if not raw:
        return []
    parts = []
    for piece in str(raw).split(","):
        item = " ".join(piece.split())
        if item:
            parts.append(item)
    return parts


def unique(items: list[str]) -> list[str]:
    seen: set[str] = set()
    out: list[str] = []
    for item in items:
        key = item.lower()
        if key not in seen:
            seen.add(key)
            out.append(item)
    return out


def grade_band(grade_col: str, grade_sort: int) -> tuple[str, int, int]:
    text = grade_col or ""
    if grade_sort == -1:
        return "Pre-K", -1, -1
    if grade_sort == 0:
        return "Kindergarten", 0, 0
    if "high school" in text.lower() or "9-12" in text:
        return "Grades 9-12", 9, 12
    return f"Grade {grade_sort}", grade_sort, grade_sort


def concept_label(raw: object) -> str:
    if not raw:
        return ""
    names: list[str] = []
    for piece in str(raw).split(","):
        piece = piece.strip()
        if not piece:
            continue
        try:
            names.append(CONCEPTS[int(piece)])
        except (TypeError, ValueError, KeyError):
            names.append(piece)
    return " ".join(names)


def fallback_summary(title: str, grade_range: str, theme: str) -> str:
    unit = f" from the {theme} unit" if theme else ""
    return (
        f"{title} — a {grade_range} lesson{unit} in the Everglades Literacy Teacher Toolkit."
    )


def js_string(value: str) -> str:
    return json.dumps(value, ensure_ascii=False)


def js_string_array(values: list[str]) -> str:
    if not values:
        return "[]"
    inner = ", ".join(js_string(v) for v in values)
    return f"[{inner}]"


def load_pdf_meta() -> dict[str, dict]:
    if not PDF_META.exists():
        return {}
    rows = json.loads(PDF_META.read_text(encoding="utf-8"))
    return {row["id"]: row for row in rows}


def concept_from_numbers(nums: list[int]) -> str:
    return " ".join(CONCEPTS[n] for n in nums if n in CONCEPTS)


def main() -> None:
    pdf_meta = load_pdf_meta()
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb["Lesson Index"]
    headers = [c.value for c in next(ws.iter_rows(min_row=5, max_row=5))]

    lessons: list[dict] = []
    for row in ws.iter_rows(min_row=6, max_row=ws.max_row, values_only=True):
        rec = dict(zip(headers, row))
        if not rec.get("Lesson ID"):
            continue
        if rec.get("Status") != "Active":
            continue

        lesson_id = str(rec["Lesson ID"])
        extracted = pdf_meta.get(lesson_id, {})
        grade_sort = int(rec["Grade Sort"])
        grade_col = rec.get("Grade") or ""
        grade_range, grade_min, grade_max = grade_band(grade_col, grade_sort)
        theme = (rec.get("Theme / Topic") or "").strip()
        title = str(rec["Title of Lesson"]).strip()
        topics = unique(
            split_list(rec.get("Topic Tags"))
            + split_list(theme)
            + extracted.get("vocab", [])
            + extracted.get("prekDomains", [])
        )
        standards = split_list(rec.get("Standards (as published)")) or extracted.get(
            "standards", []
        )
        sheet_concepts = (rec.get("Fundamental Concepts") or "").strip()
        concept = (
            concept_label(sheet_concepts)
            if sheet_concepts
            else concept_from_numbers(extracted.get("concepts") or [])
        )
        summary = (
            (rec.get("Summary of Lesson") or "").strip()
            or (extracted.get("summary") or "").strip()
            or fallback_summary(title, grade_range, theme)
        )
        pdf = (rec.get("pdfUrl") or "").strip()
        folder = (rec.get("lessonUrl") or "").strip()

        lessons.append(
            {
                "id": str(rec["Lesson ID"]),
                "title": title,
                "gradeRange": grade_range,
                "gradeMin": grade_min,
                "gradeMax": grade_max,
                "topics": topics,
                "ngsssStandards": standards,
                "fundamentalConcept": concept,
                "summary": summary,
                "lessonUrl": folder,
                "pdfUrl": pdf,
            }
        )

    blocks = []
    for les in lessons:
        blocks.append(
            "\n".join(
                [
                    "  {",
                    f"    id: {js_string(les['id'])},",
                    f"    title: {js_string(les['title'])},",
                    f"    gradeRange: {js_string(les['gradeRange'])},",
                    f"    gradeMin: {les['gradeMin']},",
                    f"    gradeMax: {les['gradeMax']},",
                    f"    topics: {js_string_array(les['topics'])},",
                    f"    ngsssStandards: {js_string_array(les['ngsssStandards'])},",
                    f"    fundamentalConcept: {js_string(les['fundamentalConcept'])},",
                    f"    summary: {js_string(les['summary'])},",
                    f"    lessonUrl: {js_string(les['lessonUrl'])},",
                    f"    pdfUrl: {js_string(les['pdfUrl'])},",
                    "  },",
                ]
            )
        )

    header = """import type { Lesson } from "../types";

// Active lessons from data/Everglades_Master_Lesson_Index.xlsx,
// with overviews/standards filled from the lesson PDFs when the sheet is blank.
// Draft / Under review rows are omitted (Legend: only Active is served).
// lessonUrl is the Google Drive folder; pdfUrl is the direct PDF download.
// Regenerate: python3 scripts/extract-pdf-metadata.py && python3 scripts/build-lessons.py
export const LESSONS: Lesson[] = [
"""
    OUT.write_text(header + "\n".join(blocks) + "\n];\n", encoding="utf-8")
    print(f"Wrote {len(lessons)} lessons to {OUT.relative_to(ROOT)}")


if __name__ == "__main__":
    main()
