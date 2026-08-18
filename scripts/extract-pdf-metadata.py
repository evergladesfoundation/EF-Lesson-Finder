#!/usr/bin/env python3
"""Pull overview, standards, concepts, and vocabulary from Teacher Toolkit PDFs.

Reads Active rows from the Master Lesson Index, downloads each public PDF
(unless already cached), and writes data/pdf-metadata.json for build-lessons.py.
"""

from __future__ import annotations

import json
import re
import ssl
from pathlib import Path
from urllib.request import Request, urlopen

import openpyxl
from pypdf import PdfReader

ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / "data" / "Everglades_Master_Lesson_Index.xlsx"
OUT = ROOT / "data" / "pdf-metadata.json"
CACHE = Path("/tmp/lesson-pdfs")
UA = "Mozilla/5.0 (compatible; EvergladesLessonFinder/1.0)"
CTX = ssl.create_default_context()

STOP_HEADINGS = [
    "Background",
    "Duration",
    "Materials",
    "Vocabulary",
    "Procedure",
    "Objective",
    "Objectives",
    "Standards",
    "Preparation",
    "Grade Level",
    "At A Glance",
    "Overview",
    "Part One",
    "Assessment",
]

VOCAB_JUNK = (
    "powerpoint",
    "copies",
    "worksheet",
    "click",
    "begin ",
    "refrain",
    "student page",
    "teacher key",
)

STANDARD_RE = re.compile(
    r"\b((?:SC|ELA|SS|HE|MAFS|MA|LAFS)\.[A-Z0-9]+(?:\.[A-Z0-9]+){1,6}(?:\.[a-z])?)\b"
)


def download_pdf(url: str, dest: Path) -> bool:
    if dest.exists() and dest.stat().st_size > 1000 and dest.read_bytes()[:4] == b"%PDF":
        return True
    req = Request(url, headers={"User-Agent": UA})
    with urlopen(req, timeout=90, context=CTX) as resp:
        data = resp.read()
    if data[:4] != b"%PDF":
        return False
    dest.write_bytes(data)
    return True


def pdf_text(path: Path, max_pages: int = 3) -> str:
    reader = PdfReader(str(path))
    return "\n".join((page.extract_text() or "") for page in reader.pages[:max_pages])


def mend_standard_dots(text: str) -> str:
    # SC.7 .E.6.6  and  SC.912.L.17 .16  and  LAFS.8.SL.1 .1
    text = re.sub(r"(?<=\d)\s+\.(?=[A-Z])", ".", text)
    text = re.sub(r"(?<=[A-Z0-9])\s*\.\s*(?=\d)", ".", text)
    text = re.sub(r"(?<=\d)\s*\.\s*(?=[a-z]\b)", ".", text)
    return text


def is_complete_code(code: str) -> bool:
    parts = code.split(".")
    if len(parts) < 3:
        return False
    last = parts[-1]
    return bool(re.fullmatch(r"\d+[a-z]?", last) or re.fullmatch(r"[a-z]", last))


def extract_standards(text: str) -> list[str]:
    found: list[str] = []
    seen: set[str] = set()
    for match in STANDARD_RE.findall(mend_standard_dots(text)):
        if match[-1].islower():
            code = match[:-1].upper() + match[-1]
        else:
            code = match.upper()
        if not is_complete_code(code) or code in seen:
            continue
        seen.add(code)
        found.append(code)
    return found


def extract_concepts(text: str) -> list[int]:
    nums: list[int] = []
    for match in re.finditer(r"Fundamental Concept\s+(\d)", text, re.I):
        n = int(match.group(1))
        if 1 <= n <= 7 and n not in nums:
            nums.append(n)
    return nums


def clean_para(value: str) -> str:
    value = re.sub(r"\s+", " ", value).strip()
    value = value.replace(" ,", ",").replace(" .", ".")
    value = value.replace("unique- ness", "uniqueness").replace("- ness", "ness")
    return value


def trim_sentences(value: str, max_len: int = 480) -> str:
    value = value.strip()
    if len(value) <= max_len:
        return value
    cut = value[:max_len]
    for sep in (". ", "! ", "? "):
        idx = cut.rfind(sep)
        if idx >= 60:
            return cut[: idx + 1]
    return cut.rsplit(" ", 1)[0] + "…"


def paragraph_after(text: str, heading: str) -> str:
    pattern = re.compile(re.escape(heading) + r"\s*:?\s*", re.I)
    match = pattern.search(text)
    if not match:
        return ""
    rest = text[match.end() :]
    stop = re.search(
        r"(?:" + "|".join(re.escape(h) for h in STOP_HEADINGS) + r")\s*:",
        rest,
        re.I,
    )
    chunk = rest[: stop.start()] if stop else rest[:900]
    return clean_para(chunk)


def usable_summary(candidate: str) -> bool:
    if len(candidate) < 40:
        return False
    if candidate.lower().startswith("students will be able to:") and len(candidate) < 80:
        return False
    return True


def extract_summary(text: str) -> str:
    for heading in ("At A Glance", "Overview"):
        candidate = paragraph_after(text, heading)
        if usable_summary(candidate):
            return trim_sentences(candidate)

    for match in re.finditer(
        r"Students will (?!be able to:)(.{40,450}?)(?:\n\n|\.|$)",
        text,
        re.S,
    ):
        candidate = clean_para("Students will " + match.group(1))
        if "able to:" in candidate[:40]:
            continue
        if len(candidate) >= 50:
            return trim_sentences(candidate)
    return ""


def extract_prek_domains(text: str) -> list[str]:
    names: list[str] = []
    for match in re.finditer(
        r"\b((?:I{1,3}|IV|VI{0,3}|IX|X)\.\s+[A-Z][A-Za-z ]{8,50})",
        text,
    ):
        name = clean_para(match.group(1))
        name = re.sub(r"\s+[A-Z]$", "", name).strip()
        # Drop trailing subsection fragments like "Through the A"
        name = re.sub(r"\s+Through the A$", " Through the Arts", name)
        if name not in names:
            names.append(name)
    return names[:6]


def extract_vocab(text: str) -> list[str]:
    match = re.search(
        r"Vocabulary\s*:(.*?)(?:Procedure|Part One|Assessment|Duration|Materials)\s*:",
        text,
        re.S | re.I,
    )
    if not match:
        return []
    terms: list[str] = []
    for term_match in re.finditer(
        r"([A-Z][A-Za-z][A-Za-z /-]{2,40})\s+[—–-]\s+",
        match.group(1),
    ):
        term = term_match.group(1).strip()
        low = term.lower()
        if any(junk in low for junk in VOCAB_JUNK):
            continue
        if len(term.split()) > 4:
            continue
        if term not in terms:
            terms.append(term)
    return terms[:8]


def load_active_rows() -> list[dict]:
    workbook = openpyxl.load_workbook(XLSX, data_only=True)
    sheet = workbook["Lesson Index"]
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=5, max_row=5))]
    rows = []
    for row in sheet.iter_rows(min_row=6, max_row=sheet.max_row, values_only=True):
        rec = dict(zip(headers, row))
        if rec.get("Lesson ID") and rec.get("Status") == "Active":
            rows.append(rec)
    return rows


def main() -> None:
    CACHE.mkdir(parents=True, exist_ok=True)
    results: list[dict] = []
    for rec in load_active_rows():
        lesson_id = str(rec["Lesson ID"])
        url = (rec.get("pdfUrl") or "").strip()
        if not url:
            print(f"{lesson_id}: no pdfUrl")
            continue
        dest = CACHE / f"{lesson_id}.pdf"
        print(f"{lesson_id}: extracting…")
        if not download_pdf(url, dest):
            print(f"{lesson_id}: download failed")
            continue
        text = pdf_text(dest)
        entry = {
            "id": lesson_id,
            "standards": extract_standards(text),
            "concepts": extract_concepts(text),
            "summary": extract_summary(text),
            "prekDomains": extract_prek_domains(text) if lesson_id.startswith("PK") else [],
            "vocab": extract_vocab(text),
        }
        results.append(entry)
        print(
            f"  standards={len(entry['standards'])} "
            f"concepts={entry['concepts']} "
            f"summary={bool(entry['summary'])}"
        )

    OUT.write_text(json.dumps(results, indent=2) + "\n", encoding="utf-8")
    missing_summary = [r["id"] for r in results if not r["summary"]]
    missing_std = [r["id"] for r in results if not r["standards"] and not r["prekDomains"]]
    print(f"\nWrote {len(results)} records to {OUT.relative_to(ROOT)}")
    print("missing summary:", missing_summary or "none")
    print("missing standards:", missing_std or "none")


if __name__ == "__main__":
    main()
