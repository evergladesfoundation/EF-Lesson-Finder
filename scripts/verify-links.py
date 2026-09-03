#!/usr/bin/env python3
"""Check Master Lesson Index Drive folders and PDF download links."""

from __future__ import annotations

import argparse
import json
import re
import ssl
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from urllib.error import HTTPError
from urllib.request import Request, urlopen

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / "data" / "Everglades_Master_Lesson_Index.xlsx"
UA = "Mozilla/5.0 (compatible; EvergladesLessonFinder/1.0)"
CTX = ssl.create_default_context()
FOLDER_FAIL = (
    "you need permission",
    "you need access",
    "request access",
    "file not found",
    "sorry, the file you have requested does not exist",
)


def fetch(url: str, max_bytes: int = 65536) -> dict:
    req = Request(url, headers={"User-Agent": UA, "Accept": "*/*"})
    try:
        with urlopen(req, timeout=25, context=CTX) as resp:
            return {
                "status": resp.status,
                "final": resp.geturl(),
                "ctype": resp.headers.get("Content-Type", ""),
                "disp": resp.headers.get("Content-Disposition", ""),
                "body": resp.read(max_bytes),
                "error": None,
            }
    except HTTPError as e:
        body = b""
        try:
            body = e.read(max_bytes)
        except Exception:
            pass
        return {
            "status": e.code,
            "final": getattr(e, "url", url) or url,
            "ctype": e.headers.get("Content-Type", "") if e.headers else "",
            "disp": e.headers.get("Content-Disposition", "") if e.headers else "",
            "body": body,
            "error": str(e),
        }
    except Exception as e:
        return {
            "status": None,
            "final": url,
            "ctype": "",
            "disp": "",
            "body": b"",
            "error": f"{type(e).__name__}: {e}",
        }


def classify_folder(url: str) -> dict:
    r = fetch(url)
    text = r["body"].decode("utf-8", "ignore").lower()
    title = ""
    m = re.search(r"<title>(.*?)</title>", r["body"].decode("utf-8", "ignore"), re.I | re.S)
    if m:
        title = re.sub(r"\s+", " ", m.group(1)).strip()
    permission = any(s in text for s in FOLDER_FAIL)
    signin = "accounts.google.com" in (r["final"] or "").lower()
    ok = r["status"] == 200 and not permission and not signin and "google" in text
    return {
        "kind": "folder",
        "verdict": "ok" if ok else "fail",
        "status": r["status"],
        "title": title,
        "error": r["error"],
        "url": url,
    }


def classify_pdf(url: str) -> dict:
    if not url:
        return {
            "kind": "pdf",
            "verdict": "missing",
            "status": None,
            "title": "",
            "error": "no pdfUrl in index",
            "url": url,
        }
    r = fetch(url, max_bytes=4096)
    text = r["body"].decode("utf-8", "ignore").lower()
    ctype = (r["ctype"] or "").lower()
    disp = r["disp"] or ""
    final = r["final"] or ""
    filename = ""
    m = re.search(r'filename\*?=(?:UTF-8\'\')?"?([^";]+)', disp, re.I)
    if m:
        filename = m.group(1)
    is_pdf = (
        "pdf" in ctype
        or "octet-stream" in ctype
        or ".pdf" in disp.lower()
        or r["body"][:5] == b"%PDF-"
        or "googleusercontent.com" in final
    )
    virus_scan = "virus scan" in text or "download anyway" in text
    permission = any(s in text for s in FOLDER_FAIL)
    ok = (is_pdf or virus_scan) and not permission
    return {
        "kind": "pdf",
        "verdict": "ok" if ok else ("missing" if not url else "fail"),
        "status": r["status"],
        "title": filename,
        "error": r["error"],
        "url": url,
    }


def load_rows() -> list[dict]:
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb["Lesson Index"]
    headers = [c.value for c in next(ws.iter_rows(min_row=5, max_row=5))]
    rows = []
    for row in ws.iter_rows(min_row=6, max_row=ws.max_row, values_only=True):
        rec = dict(zip(headers, row))
        if rec.get("Lesson ID"):
            rows.append(rec)
    return rows


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--json", action="store_true")
    args = parser.parse_args()

    rows = load_rows()
    jobs = []
    for rec in rows:
        jobs.append(("folder", rec, rec.get("lessonUrl") or "", classify_folder))
        jobs.append(("pdf", rec, rec.get("pdfUrl") or "", classify_pdf))

    results = []
    with ThreadPoolExecutor(max_workers=8) as pool:
        futs = {pool.submit(fn, url): (kind, rec) for kind, rec, url, fn in jobs}
        for fut in as_completed(futs):
            kind, rec = futs[fut]
            out = fut.result()
            out["id"] = rec["Lesson ID"]
            out["lesson_status"] = rec.get("Status")
            out["lesson_title"] = rec.get("Title of Lesson")
            results.append(out)

    results.sort(key=lambda r: (str(r["id"]), r["kind"]))
    if args.json:
        print(json.dumps(results, indent=2))
        return

    fails = [r for r in results if r["verdict"] != "ok"]
    print(f"Checked {len(results)} URLs across {len(rows)} index rows.")
    for r in results:
        mark = "OK" if r["verdict"] == "ok" else r["verdict"].upper()
        label = r.get("title") or r.get("error") or ""
        print(f"  {mark:8} {r['kind']:6} {r['id']:8}  {label}")
    print(f"\nNon-OK: {len(fails)}")
    raise SystemExit(1 if any(r["verdict"] == "fail" for r in fails) else 0)


if __name__ == "__main__":
    main()
